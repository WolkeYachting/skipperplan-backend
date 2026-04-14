"""
app.py  –  Skipperplan Backend
"""
from flask import Flask, jsonify, send_file, request
from flask_cors import CORS
import requests, os, io, re, datetime, json, base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app, origins=["https://wolkeyachting.github.io"])

JTC_BASE       = "https://api-aws.join-the-crew.com"
JTC_HEADERS    = {"Accept":"application/json, text/plain, */*","Accept-Language":"de,en-US;q=0.9,en;q=0.8","Origin":"https://join-the-crew.com","Referer":"https://join-the-crew.com/","X-Vendor":"jtc"}
ALGOLIA_HOST   = "https://m6n4q601zw-dsn.algolia.net"
ALGOLIA_APP_ID = "M6N4Q601ZW"
ALGOLIA_INDEX  = "prod_trip_dates_skipperplan_de"
HITS_PER_PAGE  = 100
GITHUB_REPO    = "wolkeyachting/Skipperplan"

JTC_USER      = os.environ.get("JTC_USER",      "")
JTC_PASSWORD  = os.environ.get("JTC_PASSWORD",  "")
APP_PASSWORD  = os.environ.get("APP_PASSWORD",  "")
GITHUB_TOKEN  = os.environ.get("GITHUB_TOKEN",  "")

STAMMSKIPPER = {
    6429:'Julian S', 13707:'Miriam M', 2430:'Gregor K', 13035:'Nicole M',
    951:'Martin S',  14111:'Leon H',   17277:'David H',  28325:'Adam B',
    7022:'Mareike W',20147:'Yanik K',  6896:'Malte W',   9724:'Luke W',
    6758:'Marc L',   347:'Mario L',    9375:'Clemens P',
}

COLUMN_ORDER = [
    'Startdatum','Enddatum','Reisename','Typ','Tage','Altersgruppe',
    'Skipper','ID Skipper','Plätze gesamt','Plätze belegt',
    'Saison','Preisspanne (€)','Land','Kontinent','Anbieter',
    'Trip ID','Trip Date ID','Yacht ID','Yachtmodell','Yachtname',
    'Bootstyp','Baujahr','Yacht Status','Skipper Status','Flotillenführer','Bewerbungen',
]
COL_WIDTHS = {
    'Startdatum':12,'Enddatum':12,'Reisename':28,'Typ':10,'Tage':7,'Altersgruppe':13,
    'Skipper':20,'ID Skipper':10,'Plätze gesamt':14,'Plätze belegt':14,'Saison':9,
    'Preisspanne (€)':15,'Land':12,'Kontinent':11,'Anbieter':10,'Trip ID':9,
    'Trip Date ID':13,'Yacht ID':10,'Yachtmodell':22,'Yachtname':18,'Bootstyp':12,
    'Baujahr':9,'Yacht Status':15,'Skipper Status':22,'Flotillenführer':15,'Bewerbungen':40,
}

COL_HEADER      = '1F4E79'
COL_ROW_A       = 'D6E4F0'
COL_ROW_B       = 'FFFFFF'
COL_NO_SKIPPER  = 'B4F7B4'
COL_CANCEL      = 'FFCCCC'
COL_UNCONFIRMED = 'FCE4D6'
COL_STAMM       = 'C6EFCE'
COL_ADVANCED    = 'FFEB9C'
COL_CHANGE      = 'FFF2CC'
COL_NEW         = 'E2EFDA'
COL_CANCEL_LOG  = 'FCE4D6'

SKIPPER_STATUS_LABELS = {
    'confirmed_by_admin_and_skipper': 'Bestätigt',
    'confirmed_by_admin':             'Vom Admin bestätigt',
    'assigned':                       'Zugeteilt (offen)',
}

# ── Hilfsfunktionen ───────────────────────────────────────────────────────────
def ts_to_date(ts):
    if not ts: return ''
    try: return datetime.datetime.fromtimestamp(int(ts), datetime.UTC).strftime('%d.%m.%Y')
    except: return ''

def ts_to_iso(ts):
    if not ts: return ''
    try: return datetime.datetime.fromtimestamp(int(ts), datetime.UTC).strftime('%Y-%m-%d')
    except: return ''

def first(lst, default=''):
    return lst[0] if lst else default

def thin_border():
    """Immer neue Instanz erstellen – niemals dasselbe Objekt an mehrere Zellen weitergeben."""
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def styled_cell(cell, bg, bold=False, center=False):
    """Wendet Stil auf eine Zelle an – erzeugt dabei immer neue Style-Objekte."""
    cell.font      = Font(name='Arial', size=11 if bold else 10, bold=bold,
                          color='FFFFFF' if bold else '000000')
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center' if (bold or center) else 'left',
                               vertical='center', wrap_text=bold)
    cell.border    = thin_border()

def days_to_weeks(days):
    return round(days / 7)

def parse_days(s):
    m = re.search(r'(\d+)', str(s or ''))
    return int(m.group(1)) if m else 7

def today_iso():
    return datetime.datetime.now(datetime.UTC).strftime('%Y-%m-%d')

# ── GitHub API ────────────────────────────────────────────────────────────────
def gh_headers():
    return {"Authorization": f"Bearer {GITHUB_TOKEN}", "Accept": "application/vnd.github+json"}

def gh_get_file(path):
    """Gibt (content_bytes, sha) zurück oder (None, None) wenn nicht vorhanden."""
    resp = requests.get(f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}",
                        headers=gh_headers(), timeout=15)
    if resp.status_code == 404:
        return None, None
    resp.raise_for_status()
    data = resp.json()
    return base64.b64decode(data['content']), data['sha']

def gh_put_file(path, content_bytes, sha, message):
    payload = {"message": message, "content": base64.b64encode(content_bytes).decode()}
    if sha:
        payload["sha"] = sha
    resp = requests.put(f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}",
                        headers=gh_headers(), json=payload, timeout=30)
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"GitHub Fehler {resp.status_code}: {resp.text[:200]}")

def load_history():
    content, _ = gh_get_file("history.json")
    if content is None:
        return {"daily_hits": [], "currently_sailing": [], "past_trips": []}
    return json.loads(content.decode('utf-8'))

def save_history(history, message="History aktualisiert"):
    _, sha = gh_get_file("history.json")
    gh_put_file("history.json", json.dumps(history, ensure_ascii=False).encode('utf-8'), sha, message)

# ── Daten laden ───────────────────────────────────────────────────────────────
def login():
    resp = requests.post(f"{JTC_BASE}/de/api/users/login",
                         json={"user": JTC_USER, "password": JTC_PASSWORD},
                         headers=JTC_HEADERS, timeout=15)
    resp.raise_for_status()
    key = (resp.json().get("api_keys") or {}).get("algolia", {}).get("skipperplan")
    if not key: raise ValueError("Algolia-Key nicht gefunden")
    return key

def fetch_algolia(api_key):
    hdrs = {"x-algolia-application-id": ALGOLIA_APP_ID, "x-algolia-api-key": api_key, "Content-Type": "application/json"}
    all_hits, page, total = [], 0, None
    while True:
        body = {"requests": [{"indexName": ALGOLIA_INDEX, "params": f"hitsPerPage={HITS_PER_PAGE}&page={page}&query="}]}
        resp = requests.post(f"{ALGOLIA_HOST}/1/indexes/*/queries", headers=hdrs, json=body, timeout=30)
        resp.raise_for_status()
        result = resp.json()["results"][0]
        hits   = result.get("hits", [])
        all_hits.extend(hits)
        if total is None: total = result.get("nbHits", 0)
        if len(all_hits) >= total or len(hits) < HITS_PER_PAGE: break
        page += 1
    seen, unique = set(), []
    for h in all_hits:
        tid = h.get("trip_date_id") or h.get("objectID")
        if tid not in seen:
            seen.add(tid); unique.append(h)
    return unique

# ── Zeilen extrahieren ────────────────────────────────────────────────────────
def extract_rows(hits):
    rows = []
    for h in hits:
        dest      = h.get('trip_destination', {}) or {}
        proposals = [p['name'] for p in h.get('yacht_skipper_proposals', []) if p.get('name')]
        start_ts  = h.get('start_date_min') or first(h.get('start_date', []))
        end_ts    = first(h.get('end_date', []))
        trip_info = {
            'Startdatum':      ts_to_date(start_ts),
            'Enddatum':        ts_to_date(end_ts),
            '_start_iso':      ts_to_iso(start_ts),
            '_end_iso':        ts_to_iso(end_ts),
            'Reisename':       dest.get('name', ''),
            'Typ':             h.get('type', ''),
            'Tage':            first(h.get('trip_days', [])),
            'Altersgruppe':    first(h.get('age_range', [])),
            'Saison':          first(h.get('season', [])),
            'Preisspanne (€)': first(h.get('price_range', [])),
            'Land':            dest.get('country', ''),
            'Kontinent':       dest.get('continent', ''),
            'Anbieter':        h.get('vendor', ''),
            'Trip ID':         h.get('trip_id', ''),
            'Trip Date ID':    h.get('trip_date_id', ''),
        }
        for i, y in enumerate(h.get('yachts', [])):
            sk            = (y.get('skipper') or {})
            sk_status_raw = sk.get('status', '')
            sk_name       = sk.get('name', '')
            yacht_status  = y.get('status', '')
            if yacht_status == 'should_be_canceled': row_color = COL_CANCEL
            elif not sk_name:                         row_color = COL_NO_SKIPPER
            elif sk_status_raw == 'assigned':         row_color = COL_UNCONFIRMED
            else:                                     row_color = None
            row = dict(trip_info)
            row.update({
                'Skipper':         sk_name if sk_name else '⚠ Kein Skipper',
                'ID Skipper':      sk.get('id', ''),
                'Plätze gesamt':   y.get('places', ''),
                'Plätze belegt':   y.get('occupied_places', ''),
                'Yacht ID':        y.get('id', ''),
                'Yachtmodell':     y.get('accomodation_details_name', ''),
                'Yachtname':       y.get('yacht_name', ''),
                'Bootstyp':        y.get('accomodation_details_type', ''),
                'Baujahr':         y.get('yacht_year', ''),
                'Yacht Status':    'Zu stornieren' if yacht_status == 'should_be_canceled' else 'Bestätigt',
                'Skipper Status':  SKIPPER_STATUS_LABELS.get(sk_status_raw, sk_status_raw or '—'),
                'Flotillenführer': 'Ja' if sk.get('is_flotilla_leader') else 'Nein',
                'Bewerbungen':     ', '.join(proposals) if i == 0 else '',
                '_color':          row_color,
                '_yacht_key':      f"{h.get('trip_date_id')}_{y.get('id')}",
                '_proposals':      proposals if i == 0 else [],
            })
            rows.append(row)
    return rows

def build_skipper_data(hits):
    from collections import defaultdict
    data = defaultdict(lambda: {'name':'','törns':0,'total_weeks':0,'summer_weeks':0})
    for h in hits:
        weeks  = days_to_weeks(parse_days(first(h.get('trip_days', []))))
        season = (first(h.get('season', '')) or '').lower()
        for y in h.get('yachts', []):
            sk = y.get('skipper') or {}
            if not sk.get('id') or not sk.get('name'): continue
            d = data[sk['id']]
            d['name'] = sk['name']
            d['törns'] += 1
            d['total_weeks'] += weeks
            if season == 'summer': d['summer_weeks'] += weeks
    return sorted(data.items(), key=lambda x: -x[1]['törns'])

# ── Changelog ────────────────────────────────────────────────────────────────
def detect_changes(new_hits, prev_hits, today):
    """Vergleicht neue mit vorherigen Hits und gibt Änderungen zurück."""
    changes = []

    # Index aufbauen: yacht_key → Daten
    def index_hits(hits):
        idx = {}
        for h in hits:
            dest      = h.get('trip_destination', {}) or {}
            start_ts  = h.get('start_date_min') or first(h.get('start_date', []))
            end_ts    = first(h.get('end_date', []))
            proposals = [p['name'] for p in h.get('yacht_skipper_proposals', []) if p.get('name')]
            for y in h.get('yachts', []):
                key = f"{h.get('trip_date_id')}_{y.get('id')}"
                sk  = y.get('skipper') or {}
                idx[key] = {
                    'trip_date_id':  h.get('trip_date_id'),
                    'yacht_id':      y.get('id'),
                    'reisename':     dest.get('name', ''),
                    'startdatum':    ts_to_date(start_ts),
                    'start_iso':     ts_to_iso(start_ts),
                    'enddatum':      ts_to_date(end_ts),
                    'end_iso':       ts_to_iso(end_ts),
                    'yachtmodell':   y.get('accomodation_details_name', ''),
                    'yachtname':     y.get('yacht_name', ''),
                    'skipper_name':  sk.get('name', ''),
                    'skipper_id':    sk.get('id', ''),
                    'yacht_status':  y.get('status', ''),
                    'proposals':     proposals,
                }
        return idx

    new_idx  = index_hits(new_hits)
    prev_idx = index_hits(prev_hits)

    # Neue Boote
    for key, nd in new_idx.items():
        if key not in prev_idx:
            changes.append({
                'typ': 'Neues Boot',
                **nd,
                'skipper_alt': '',
            })

    # Storniert oder geändert
    for key, pd in prev_idx.items():
        if key not in new_idx:
            if pd['start_iso'] > today:
                # Startdatum in Zukunft → storniert
                changes.append({'typ': 'Storniert', **pd, 'skipper_alt': pd['skipper_name']})
            # Wenn start_iso <= today → Boot ist losgefahren, kein Changelog-Eintrag
        else:
            nd = new_idx[key]
            row_changes = {}

            # Skipper geändert
            if pd['skipper_name'] != nd['skipper_name']:
                row_changes['skipper_alt'] = pd['skipper_name'] or '(Kein Skipper)'
                row_changes['typ'] = 'Skipper geändert'

            # Neue Bewerbungen
            new_proposals = [p for p in nd['proposals'] if p not in pd['proposals']]
            if new_proposals:
                row_changes['neue_bewerbungen'] = ', '.join(new_proposals)
                if 'typ' not in row_changes:
                    row_changes['typ'] = 'Neue Bewerbung(en)'

            if row_changes:
                changes.append({**nd, **row_changes,
                                 'skipper_alt': row_changes.get('skipper_alt', '')})

    return changes

# ── Fahrstatus aktualisieren ──────────────────────────────────────────────────
def update_sailing_status(history, new_hits, prev_hits, today):
    """Aktualisiert currently_sailing und past_trips."""
    new_idx  = {f"{h.get('trip_date_id')}_{y.get('id')}": True
                for h in new_hits for y in h.get('yachts', [])}

    # Aus prev_hits: Boote die verschwunden sind und heute fahren → currently_sailing
    existing_keys = {r['_yacht_key'] for r in history.get('currently_sailing', [])}
    for h in prev_hits:
        dest     = h.get('trip_destination', {}) or {}
        start_ts = h.get('start_date_min') or first(h.get('start_date', []))
        end_ts   = first(h.get('end_date', []))
        start_iso = ts_to_iso(start_ts)
        end_iso   = ts_to_iso(end_ts)
        proposals = [p['name'] for p in h.get('yacht_skipper_proposals', []) if p.get('name')]
        for i, y in enumerate(h.get('yachts', [])):
            key = f"{h.get('trip_date_id')}_{y.get('id')}"
            if key not in new_idx and key not in existing_keys:
                if start_iso and start_iso <= today and end_iso >= today:
                    sk = y.get('skipper') or {}
                    row = {
                        'Startdatum':      ts_to_date(start_ts),
                        'Enddatum':        ts_to_date(end_ts),
                        '_start_iso':      start_iso,
                        '_end_iso':        end_iso,
                        '_yacht_key':      key,
                        'Reisename':       dest.get('name', ''),
                        'Typ':             h.get('type', ''),
                        'Tage':            first(h.get('trip_days', [])),
                        'Altersgruppe':    first(h.get('age_range', [])),
                        'Saison':          first(h.get('season', [])),
                        'Preisspanne (€)': first(h.get('price_range', [])),
                        'Land':            dest.get('country', ''),
                        'Kontinent':       dest.get('continent', ''),
                        'Anbieter':        h.get('vendor', ''),
                        'Trip ID':         h.get('trip_id', ''),
                        'Trip Date ID':    h.get('trip_date_id', ''),
                        'Skipper':         sk.get('name', '') or '—',
                        'ID Skipper':      sk.get('id', ''),
                        'Plätze gesamt':   y.get('places', ''),
                        'Plätze belegt':   y.get('occupied_places', ''),
                        'Yacht ID':        y.get('id', ''),
                        'Yachtmodell':     y.get('accomodation_details_name', ''),
                        'Yachtname':       y.get('yacht_name', ''),
                        'Bootstyp':        y.get('accomodation_details_type', ''),
                        'Baujahr':         y.get('yacht_year', ''),
                        'Yacht Status':    'Bestätigt',
                        'Skipper Status':  SKIPPER_STATUS_LABELS.get(sk.get('status',''), sk.get('status','') or '—'),
                        'Flotillenführer': 'Ja' if sk.get('is_flotilla_leader') else 'Nein',
                        'Bewerbungen':     ', '.join(proposals) if i == 0 else '',
                    }
                    history['currently_sailing'].append(row)

    # currently_sailing → past_trips wenn Enddatum überschritten
    still_sailing, ended = [], []
    cutoff_delete = (datetime.datetime.now(datetime.UTC) - datetime.timedelta(weeks=4)).strftime('%Y-%m-%d')
    for row in history.get('currently_sailing', []):
        end_iso = row.get('_end_iso', '')
        if end_iso and end_iso < today:
            ended.append(row)
        else:
            still_sailing.append(row)

    history['currently_sailing'] = still_sailing

    # Vergangene hinzufügen (Duplikate vermeiden)
    past_keys = {r['_yacht_key'] for r in history.get('past_trips', [])}
    for row in ended:
        if row['_yacht_key'] not in past_keys:
            history['past_trips'].append(row)

    # Vergangene löschen wenn > 4 Wochen her
    history['past_trips'] = [r for r in history.get('past_trips', [])
                              if r.get('_end_iso', '') >= cutoff_delete]

    return history

# ── Excel bauen ───────────────────────────────────────────────────────────────
def apply_header(ws, cols, widths=None):
    ws.row_dimensions[1].height = 32
    for c, col in enumerate(cols, 1):
        cell = ws.cell(1, c, col)
        styled_cell(cell, COL_HEADER, bold=True)
        if widths:
            ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 14)

def data_cell(cell, bg, bold=False):
    styled_cell(cell, bg, bold=bold)

def add_data_sheet(wb, title, rows, freeze=True):
    ws = wb.create_sheet(title)
    apply_header(ws, COLUMN_ORDER, COL_WIDTHS)
    fill_a = PatternFill('solid', start_color=COL_ROW_A)
    fill_b = PatternFill('solid', start_color=COL_ROW_B)
    for r, row in enumerate(rows, 2):
        override = row.get('_color')
        rf = PatternFill('solid', start_color=override) if override else (fill_a if r%2==0 else fill_b)
        for c, col in enumerate(COLUMN_ORDER, 1):
            cell = ws.cell(r, c, row.get(col, ''))
            data_cell(cell, override or (COL_ROW_A if r%2==0 else COL_ROW_B))
    if freeze:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_ORDER))}1"
    return ws

def build_excel(rows, skipper_data, changelog=None, currently_sailing=None, past_trips=None):
    wb = Workbook()

    # ── Segelreisen ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Segelreisen"
    apply_header(ws, COLUMN_ORDER, COL_WIDTHS)
    for r, row in enumerate(rows, 2):
        override = row.get('_color')
        bg = override or (COL_ROW_A if r%2==0 else COL_ROW_B)
        for c, col in enumerate(COLUMN_ORDER, 1):
            cell = ws.cell(r, c, row.get(col, ''))
            styled_cell(cell, bg)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_ORDER))}1"
    # Legende
    lc = len(COLUMN_ORDER) + 2
    ws.cell(1, lc, 'Legende').font = Font(name='Arial', bold=True, size=11)
    for i, (color, label) in enumerate([(COL_NO_SKIPPER,'⚠ Kein Skipper'),(COL_CANCEL,'🚫 Zu stornieren'),(COL_UNCONFIRMED,'⏳ Nicht bestätigt')], 2):
        cell = ws.cell(i, lc, label)
        cell.fill   = PatternFill('solid', start_color=color)
        cell.font   = Font(name='Arial', size=10)
        cell.border = thin_border()
    ws.column_dimensions[get_column_letter(lc)].width = 26

    # ── Aktuell fahrende Boote ─────────────────────────────────────────────────
    ws_sail = wb.create_sheet("Aktuell fahrend")
    apply_header(ws_sail, COLUMN_ORDER, COL_WIDTHS)
    for r, row in enumerate(sorted(currently_sailing or [], key=lambda x: x.get('_start_iso','')), 2):
        bg = COL_ROW_A if r%2==0 else COL_ROW_B
        for c, col in enumerate(COLUMN_ORDER, 1):
            styled_cell(ws_sail.cell(r, c, row.get(col, '')), bg)
    ws_sail.freeze_panes = 'A2'
    ws_sail.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_ORDER))}1"

    # ── Vergangene Törns ───────────────────────────────────────────────────────
    ws_past = wb.create_sheet("Vergangene Törns")
    apply_header(ws_past, COLUMN_ORDER, COL_WIDTHS)
    for r, row in enumerate(sorted(past_trips or [], key=lambda x: x.get('_end_iso',''), reverse=True), 2):
        bg = COL_ROW_A if r%2==0 else COL_ROW_B
        for c, col in enumerate(COLUMN_ORDER, 1):
            styled_cell(ws_past.cell(r, c, row.get(col, '')), bg)
    ws_past.freeze_panes = 'A2'
    ws_past.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_ORDER))}1"

    # ── Changelog ──────────────────────────────────────────────────────────────
    if changelog is not None:
        ws_cl = wb.create_sheet("Changelog")
        cl_cols = ['Typ','Reisename','Startdatum','Enddatum','Yachtmodell','Yachtname',
                   'Skipper (neu)','Skipper (alt)','Neue Bewerbungen']
        cl_widths = {'Typ':18,'Reisename':28,'Startdatum':12,'Enddatum':12,
                     'Yachtmodell':22,'Yachtname':18,'Skipper (neu)':20,
                     'Skipper (alt)':20,'Neue Bewerbungen':40}
        apply_header(ws_cl, cl_cols, cl_widths)
        color_map = {'Skipper geändert': COL_CHANGE, 'Neues Boot': COL_NEW,
                     'Storniert': COL_CANCEL_LOG, 'Neue Bewerbung(en)': COL_ROW_A}
        for r, ch in enumerate(changelog, 2):
            bg = color_map.get(ch.get('typ',''), COL_ROW_B)
            vals = [ch.get('typ',''), ch.get('reisename',''), ch.get('startdatum',''),
                    ch.get('enddatum',''), ch.get('yachtmodell',''), ch.get('yachtname',''),
                    ch.get('skipper_name',''), ch.get('skipper_alt',''), ch.get('neue_bewerbungen','')]
            for c, val in enumerate(vals, 1):
                styled_cell(ws_cl.cell(r, c, val), bg)
        ws_cl.freeze_panes = 'A2'
        ws_cl.auto_filter.ref = f"A1:{get_column_letter(len(cl_cols))}1"

    # ── Skipper ────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Skipper")
    sk_cols = ['Skipper','ID','Törns','Wochen','Status']
    sk_widths = {'Skipper':24,'ID':10,'Törns':8,'Wochen':9,'Status':16}
    apply_header(ws2, sk_cols, sk_widths)
    status_colors = {'Stammskipper': COL_STAMM, 'Advanced': COL_ADVANCED, 'Hobby': COL_ROW_B}
    for r, (sid, d) in enumerate(skipper_data, 2):
        status = 'Stammskipper' if sid in STAMMSKIPPER else ('Advanced' if d['summer_weeks'] > 5 else 'Hobby')
        vals   = [d['name'], sid, d['törns'], d['total_weeks'], status]
        default_bg = COL_ROW_A if r%2==0 else COL_ROW_B
        for c, val in enumerate(vals, 1):
            bg = status_colors[status] if c==5 else default_bg
            cell = ws2.cell(r, c, val)
            styled_cell(cell, bg, center=(c in (2,3,4)))
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = 'A1:E1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Endpunkte ─────────────────────────────────────────────────────────────────
def check_password():
    provided = request.args.get("password", "")
    if APP_PASSWORD and provided != APP_PASSWORD:
        return False
    return True

@app.route("/check", methods=["GET"])
def check():
    if not check_password():
        return jsonify({"ok": False}), 401
    return jsonify({"ok": True})

@app.route("/refresh", methods=["GET"])
def refresh():
    """Daten laden, Excel erstellen, ins GitHub-Repo committen."""
    source = request.args.get("source", "manual")  # "daily" oder "manual"
    try:
        now_str  = datetime.datetime.now(datetime.UTC).strftime('%d.%m.%Y %H:%M UTC')
        today    = today_iso()

        # Daten holen
        api_key  = login()
        new_hits = fetch_algolia(api_key)

        # History laden
        history  = load_history()
        prev_hits = history.get("daily_hits", [])

        # Changelog (nur wenn vorherige Daten vorhanden)
        changelog = detect_changes(new_hits, prev_hits, today) if prev_hits else []

        # Fahrstatus aktualisieren
        history = update_sailing_status(history, new_hits, prev_hits, today)

        # Bei täglichem Lauf: daily_hits aktualisieren
        if source == "daily":
            history["daily_hits"] = new_hits

        # Excel bauen
        rows         = extract_rows(new_hits)
        skipper_data = build_skipper_data(new_hits)
        buf          = build_excel(rows, skipper_data, changelog,
                                   history['currently_sailing'], history['past_trips'])
        excel_bytes  = buf.read()

        # History speichern
        save_history(history, f"History aktualisiert: {now_str}")

        # skipperplan.xlsx speichern (immer)
        _, sha_main = gh_get_file("skipperplan.xlsx")
        gh_put_file("skipperplan.xlsx", excel_bytes, sha_main,
                    f"Skipperplan aktualisiert: {now_str}")

        # skipperplan_daily.xlsx nur beim täglichen Lauf speichern
        if source == "daily":
            _, sha_daily = gh_get_file("skipperplan_daily.xlsx")
            gh_put_file("skipperplan_daily.xlsx", excel_bytes, sha_daily,
                        f"Skipperplan täglich: {now_str}")

        return jsonify({
            "ok":       True,
            "updated":  now_str,
            "termine":  len(new_hits),
            "yachten":  len(rows),
            "changes":  len(changelog),
            "sailing":  len(history['currently_sailing']),
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()[-500:]}), 500


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
